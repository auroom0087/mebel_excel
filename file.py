import os.path
import openpyxl

def txt_to_tuple(file_dir:str) -> tuple():
    result = []
    f = open(file_dir, 'r')
    for line in f:
        line=line[:-1]
        array=line.split()
        result.append((array[0],int(array[1])))
    return result

def excel_to_tuple(file_dir:str) -> tuple():
    result = []
    file = openpyxl.load_workbook(file_dir, data_only=True)
    sheet = file['Лист2']
    for i in range(2, sheet.max_row):
        row = []
        for j in range(1, sheet.max_column):
            row.append(sheet[i][j].value)
        result.append(tuple(row))
    return result